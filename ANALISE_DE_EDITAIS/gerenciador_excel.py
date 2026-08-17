"""
V4 - Gerenciador de Excel

Cada edital processado vira uma linha na planilha. Se o arquivo ainda
não existe, cria com o cabeçalho; se já existe, só adiciona a linha no
fim — assim dá pra rodar o programa várias vezes (um edital de cada vez,
ou vários em lote) sem perder o que já foi processado antes.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

COLUNAS = [
    "Nº Edital",
    "Órgão",
    "Cidade",
    "Estado",
    "Objeto",
    "Data da Disputa",
    "Horário da Disputa",
    "Modo de Disputa",
    "Quantitativo (valor)",
    "Quantitativo (unidade)",
    "Valor Estimado Total (R$)",
    "Prazo do Contrato (meses)",
]

LARGURAS_COLUNAS = [14, 30, 16, 8, 45, 14, 12, 16, 14, 18, 20, 14]


class GerenciadorExcel:
    """Cria/atualiza a planilha de editais analisados."""

    def __init__(self, caminho_planilha: str):
        self.caminho = Path(caminho_planilha)

        if self.caminho.exists():
            self.wb = load_workbook(self.caminho)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Editais"
            self._escrever_cabecalho()

    def _escrever_cabecalho(self) -> None:
        for col_idx, titulo in enumerate(COLUNAS, start=1):
            celula = self.ws.cell(row=1, column=col_idx, value=titulo)
            celula.font = Font(name="Calibri", bold=True)
            celula.alignment = Alignment(wrap_text=True, vertical="center")

        for col_idx, largura in enumerate(LARGURAS_COLUNAS, start=1):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = largura

        self.ws.freeze_panes = "A2"  # cabeçalho fixo ao rolar

    def adicionar_edital(self, dados: dict) -> None:
        """
        `dados` pode vir da junção do ExtratorCamposFaceis (V2) com o
        ExtratorIA (V3). Chaves que não existirem viram célula vazia.
        """
        linha = [
            dados.get("numero_edital", ""),
            dados.get("orgao", ""),
            dados.get("cidade", ""),
            dados.get("estado", ""),
            dados.get("objeto", ""),
            dados.get("data_disputa", ""),
            dados.get("horario_disputa", ""),
            dados.get("modo_disputa", ""),
            dados.get("quantitativo_valor", ""),
            dados.get("quantitativo_unidade", ""),
            dados.get("valor_estimado_total", ""),
            dados.get("prazo_contrato_meses", ""),
        ]

        proxima_linha = self.ws.max_row + 1
        for col_idx, valor in enumerate(linha, start=1):
            celula = self.ws.cell(row=proxima_linha, column=col_idx, value=valor)
            celula.font = Font(name="Calibri")

        # Coluna 11 = Valor Estimado Total -> formato de moeda
        celula_valor = self.ws.cell(row=proxima_linha, column=11)
        if isinstance(celula_valor.value, (int, float)):
            celula_valor.number_format = "R$ #,##0.00"

    def salvar(self) -> None:
        self.wb.save(self.caminho)


if __name__ == "__main__":
    import sys

    from extrator_campos import ExtratorCamposFaceis
    from extrator_ia import ExtratorIA
    from leitor_pdf import LeitorPDF

    caminho_pdf = sys.argv[1] if len(sys.argv) > 1 else "edital.pdf"
    caminho_planilha = sys.argv[2] if len(sys.argv) > 2 else "editais.xlsx"

    leitor = LeitorPDF(caminho_pdf)
    leitor.extrair()

    campos_faceis = ExtratorCamposFaceis(leitor.texto_completo).extrair_tudo()
    campos_ia = ExtratorIA(leitor.texto_completo).extrair()  # precisa de ANTHROPIC_API_KEY

    dados = {**campos_faceis, **campos_ia}

    gerenciador = GerenciadorExcel(caminho_planilha)
    gerenciador.adicionar_edital(dados)
    gerenciador.salvar()

    print(f"Edital {dados.get('numero_edital')} adicionado em {caminho_planilha}")
